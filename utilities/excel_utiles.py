import os.path

import openpyxl
from openpyxl import Workbook





class ExcelUtiles:

    @staticmethod
    def write_to_cell(file,sheet_name, row_number,column_number,data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        sheet.cell(row_number,column_number).value = data
        workbook.save(filename=file)

    @staticmethod
    def read_from_cell(file, sheet_name, row_number, column_number):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row_number, column_number).value







